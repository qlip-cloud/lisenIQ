import frappe
import json
import base64
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from frappe import _
from liseniq.utils.constants import WEB_FORM_CLIENT_SCRIPT, WEB_FORM_CUSTOM_CSS
from liseniq.utils.api_survey import generate_public_link_for_survey
from liseniq.utils.login_util import global_website_context

def get_context(context):

    if frappe.session.user == "Guest":
        frappe.throw(_("Cliente aún no ha sido registrado. Por favor comunique al Administrador."), frappe.PermissionError)

    context = global_website_context(context)

    # Configuración base de la página
    context.no_cache = 1
    context.page_title = _("Crear Medición")
    context.no_breadcrumbs = True
    context.is_navbar_custom = True

    measurement_name = frappe.request.args.get('name')
    context.is_edit_mode = bool(measurement_name)
    context.page_title = _("Editar Medición") if context.is_edit_mode else _("Crear Medición")
    context.measurement_data_json = "null"

    # Preparar roles de liderazgo
    try:
        catalog = frappe.db.get_value("qp_IQ_Catalog", {"ca_mnemonico": "measurement_roles"}, "name")
        if catalog:
            context.leadership_roles = frappe.get_all(
                "qp_IQ_CatalogOptions",
                filters={"co_catalog": catalog, "co_is_active": 1},
                fields=["name", "co_label"],
                order_by="co_sort_order asc"
            )
        else:
            context.leadership_roles = []
    except Exception:
        context.leadership_roles = []

    # Modo edición: cargar datos existentes
    if context.is_edit_mode:
        try:
            doc = frappe.get_doc("qp_IQ_Survey", measurement_name)
            
            # Construir preguntas
            questions = []
            for q_link in (doc.su_questions or []):
                q_doc = frappe.get_doc("qp_IQ_Question", q_link.sq_question)
                
                t_data = frappe.db.get_value("qp_IQ_QuestionType", q_doc.qn_type, ["qnt_type_name", "qnt_mnemonico"], as_dict=True) if q_doc.qn_type else None
                type_name = t_data.qnt_type_name if t_data else "No definido"

                demo_title = frappe.db.get_value("qp_IQ_DemographicType", q_doc.qn_demographic, "dt_title") if q_doc.qn_demographic else None
                options = [opt.qo_option_text for opt in (q_doc.qn_response_options or [])]
                questions.append({
                    "id": q_doc.name,
                    "text": q_doc.qn_statement,
                    "type": q_doc.qn_type,
                    "typeName": type_name,
                    "demographic": demo_title,
                    "options": options,
                    "nps_min": q_doc.qn_nps_min,
                    "nps_max": q_doc.qn_nps_max,
                    "qp_others": q_doc.get("qp_others", 0),
                    "qp_none_above": q_doc.get("qp_none_above", 0)
                })

            # Datos de participantes
            recipients_count = frappe.db.count("qp_IQ_SurveyRecipient", {"sr_survey": doc.name})
            
            if doc.su_custom_generate_public_link and recipients_count == 0 and doc.su_is_anonymous:
                 survey_type = "anonymous_link"
            else:
                 survey_type = "selected" if recipients_count > 0 else "all"
                 
            response_type = "anonymous" if doc.su_is_anonymous else "identified"

            measurement_data = {
                "name": doc.su_name,
                "isLeadership": doc.get("su_is_leadership", 0),
                "startDate": doc.su_start_date,
                "endDate": doc.su_end_date,
                "timezone": doc.su_timezone,
                "reminders": {
                    "send": True if doc.su_send_reminders else False,
                    "frequency": doc.su_reminder_frequency,
                    "max": doc.su_reminder_max
                },
                "su_invitation_subject": doc.su_invitation_subject,
                "su_invitation_body": doc.su_invitation_body,
                "su_reminder_subject": doc.su_reminder_subject,
                "su_reminder_body": doc.su_reminder_body,
                "su_default_notif": doc.su_default_notif,
                "questions": questions,
            }

            if doc.get("su_is_leadership"):
                networks_dict = {}
                if recipients_count > 0:
                    recipients = frappe.get_all("qp_IQ_SurveyRecipient", filters={"sr_survey": doc.name}, fields=["name", "sr_contact", "sr_evaluating_to", "sr_evaluation_role"])
                    
                    contact_ids = list(set([r.sr_contact for r in recipients if r.sr_contact] + [r.sr_evaluating_to for r in recipients if r.sr_evaluating_to]))
                    contact_map = {}
                    if contact_ids:
                        contacts_data = frappe.get_all("Contact", filters={"name": ["in", contact_ids]}, fields=["name", "first_name", "last_name"])
                        for c in contacts_data:
                            contact_map[c.name] = f"{(c.first_name or '').strip()} {(c.last_name or '').strip()}".strip()

                    for r in recipients:
                        leader_id = r.sr_contact
                        evaluator_id = r.sr_evaluating_to
                        
                        # Fallback por seguridad
                        if not evaluator_id:
                            evaluator_id = leader_id

                        if leader_id not in networks_dict:
                            networks_dict[leader_id] = {
                                "leader": {"id": leader_id, "name_display": contact_map.get(leader_id, leader_id)},
                                "evaluators": []
                            }
                        
                        is_auto = (leader_id == evaluator_id)
                        role_label = r.sr_evaluation_role if r.sr_evaluation_role else ("Autoevaluación" if is_auto else "")
                        
                        networks_dict[leader_id]["evaluators"].append({
                            "id": evaluator_id,
                            "name": contact_map.get(evaluator_id, evaluator_id),
                            "role": role_label,
                            "role_id": r.sr_evaluation_role, 
                            "isAuto": is_auto
                        })
                measurement_data["leadershipNetwork"] = list(networks_dict.values())
            else:
                contacts_headers = ["Nombre"]
                contacts_list = []
                if recipients_count > 0:
                    recipient_rows = frappe.get_all("qp_IQ_SurveyRecipient", filters={"sr_survey": doc.name}, fields=["sr_contact"])
                    contact_names = [r.sr_contact for r in recipient_rows if r.sr_contact]
                    if contact_names:
                        contact_docs = frappe.get_all("Contact", filters={"name": ["in", contact_names]}, fields=["name", "first_name", "last_name"])
                        for c in contact_docs:
                            contacts_list.append({
                                "name": c.name,
                                "Nombre": f"{(c.first_name or '').strip()} {(c.last_name or '').strip()}".strip()
                            })
                measurement_data["contacts"] = {
                    "surveyType": survey_type,
                    "responseType": response_type,
                    "contactCount": recipients_count,
                    "headers": contacts_headers,
                    "list": contacts_list
                }

            context.measurement_data_json = frappe.as_json(measurement_data)
        except Exception as e:
            frappe.log_error(f"Error cargando medición para editar: {e}", "new_measurement.get_context")
            frappe.throw(_("Medición no encontrada."), frappe.DoesNotExistError)

    try:
        context.contact_demographics = frappe.get_all(
            "qp_IQ_DemographicType",
            filters={"dt_object_type": "Contacto"},
            fields=["name", "dt_title"],
            order_by="dt_title"
        )
    except frappe.DoesNotExistError:
        context.contact_demographics = []

    try:
        context.question_types = frappe.get_all(
            "qp_IQ_QuestionType",
            filters={
                "qnt_mnemonico": ["in", ["text_area", "text_short", "check_group", "scale_likert", "scale_emoji", "score_nps", "radio_group"]]
            },
            fields=["name", "qnt_type_name", "qnt_mnemonico"],
            order_by="qnt_type_name"
        )
    except frappe.DoesNotExistError:
        context.question_types = []

    try:
        meta = frappe.get_meta("qp_IQ_Survey")
        timezone_field = meta.get_field("su_timezone")
        if timezone_field and timezone_field.options:
            context.timezones = timezone_field.options.split('\n')
        else:
            context.timezones = []
    except frappe.DoesNotExistError:
        context.timezones = []

    template_name = frappe.request.args.get('template')
    template_is_leadership = False
    if template_name:
        try:
            # Identificar si la plantilla corresponde a Liderazgo
            cat_id = frappe.db.get_value("qp_IQ_Template", template_name, "tp_category")
            if cat_id:
                cat_name = frappe.db.get_value("qp_IQ_QuestionCategory", cat_id, "qnc_category")
                if cat_name == "Liderazgo":
                    template_is_leadership = True

            questions = frappe.call(
                'liseniq.www.iq-templates.index.get_questions_from_template',
                template_name=template_name
            )
            context.preloaded_questions_json = frappe.as_json(questions)
        except Exception as e:
            frappe.log_error(f"No se pudieron cargar las preguntas de la plantilla {template_name}: {e}", "Error en new_measurement.py")
            context.preloaded_questions_json = "[]"
    
    context.template_is_leadership = template_is_leadership
    context.template_name = template_name or None
    if template_name:
        frappe.cache().set_value(f"measurement_template:{frappe.session.user}", template_name)
        frappe.local.template_name = template_name

    context.update({
        "is_navbar_custom": True,
        "no_cache": 1
    })

    return context

@frappe.whitelist()
def search_company_contacts(search_term=""):
    user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")
    
    params = {}
    company_filter = ""
    
    if user_company:
        company_filter = "AND c.custom_company = %(company)s"
        params["company"] = user_company
        
    search_filter = ""
    if search_term:
        search_filter = """
            AND (
                c.first_name LIKE %(st)s OR 
                c.last_name LIKE %(st)s OR 
                CONCAT(IFNULL(c.first_name, ''), ' ', IFNULL(c.last_name, '')) LIKE %(st)s OR
                c.custom_document_number LIKE %(st)s OR
                ce.email_id LIKE %(st)s
            )
        """
        params["st"] = f"%{search_term}%"
        
    query = f"""
        SELECT DISTINCT
            c.name, c.first_name, c.last_name, c.custom_document_number, ce.email_id
        FROM
            `tabContact` c
        LEFT JOIN
            `tabContact Email` ce ON ce.parent = c.name AND ce.parenttype = 'Contact' AND ce.is_primary = 1
        WHERE
            c.status IN ('Enabled', 'Passive')
            AND c.custom_is_liseniq_contact = 1
            AND c.custom_is_deleted = 0
            AND c.custom_status != 'Inactivo'
            {company_filter}
            {search_filter}
        LIMIT 50
    """
    
    contacts = frappe.db.sql(query, params, as_dict=True)
    
    res = []
    for c in contacts:
        full_name = f"{c.first_name or ''} {c.last_name or ''}".strip()
        res.append({
            "name": c.name, 
            "full_name": full_name, 
            "email": c.email_id or "",
            "dni": c.custom_document_number or ""
        })
        
    return res

@frappe.whitelist()
def check_measurement_name(name, exclude_doc=None, only_open=False):
    user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")
    if not user_company:
        exists = frappe.db.exists("qp_IQ_Survey", {"su_name": name})
        return {"exists": bool(exists)}

    conditions = [
        ["qp_IQ_Survey", "su_owner", "=", user_company],
        ["qp_IQ_Survey", "su_name", "=", name],
    ]
    if exclude_doc:
        conditions.append(["qp_IQ_Survey", "name", "!=", exclude_doc])

    if only_open:
        open_status_docs = frappe.get_all(
            "qp_IQ_SurveyStatus",
            filters={"se_status": ["in", ["Programada", "En Progreso", "Borrador"]]},
            fields=["name"]
        )
        open_status_ids = [d.name for d in open_status_docs]
        if open_status_ids:
            conditions.append(["qp_IQ_Survey", "su_status", "in", open_status_ids])

    rows = frappe.get_all("qp_IQ_Survey", filters=conditions, fields=["name"], limit_page_length=1)
    return {"exists": bool(rows)}

@frappe.whitelist()
def get_demographic_values_for_contacts(demographic_type):
    if not demographic_type:
        return frappe._dict({"values": [], "color": None})

    color = frappe.db.get_value("qp_IQ_DemographicType", demographic_type, "dt_tag_color")

    values = frappe.get_all(
        "qp_IQ_ContactAdditionalDetail",
        filters={"cad_demographic_type": demographic_type},
        fields=["cad_value"],
        distinct=True,
        order_by="cad_value"
    )
    
    return frappe._dict({
        "values": [d.get("cad_value") for d in values if d.get("cad_value")],
        "color": color
    })

@frappe.whitelist()
def get_filtered_contacts_count(filters='[]'):
    filters = json.loads(filters)
    user_contact_name = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "name")
    user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")

    base_filters = [
        ["Contact", "status", "in", ["Enabled", "Passive"]],
        ["Contact", "custom_is_liseniq_contact", "=", 1],
        ["Contact", "custom_is_deleted", "=", 0],
        ["Contact", "custom_status", "!=", "Inactivo"]
    ]
    if user_company:
        base_filters.append(["Contact", "custom_company", "=", user_company])

    if user_contact_name:
        base_filters.append(["Contact", "name", "!=", user_contact_name])

    if not filters:
        all_contacts = frappe.get_list(
            "Contact", 
            filters=base_filters, 
            fields=["name", "first_name", "last_name"],
            ignore_permissions=True
        )
        contacts_for_modal = [{"name": c.name, "Nombre": f"{c.first_name} {c.last_name or ''}".strip()} for c in all_contacts]
        return {
            "count": len(all_contacts),
            "headers": ["Nombre"],
            "contacts": contacts_for_modal
        }

    where_conditions = []
    params = []
    demographic_ids = []
    
    for f in filters:
        demographic_type = f.get("demographic_type")
        values = f.get("values")
        if demographic_type and values:
            demographic_ids.append(demographic_type)
            value_placeholders = ','.join(['%s'] * len(values))
            where_conditions.append(f"(cad_demographic_type = %s AND cad_value IN ({value_placeholders}))")
            params.append(demographic_type)
            params.extend(values)

    if not where_conditions:
        return {"count": 0, "headers": ["Nombre"], "contacts": []}

    query = f"""
        SELECT parent
        FROM `tabqp_IQ_ContactAdditionalDetail`
        WHERE {' OR '.join(where_conditions)}
        GROUP BY parent
        HAVING COUNT(DISTINCT cad_demographic_type) = %s
    """
    params.append(len(filters))

    matching_contact_names = [row[0] for row in frappe.db.sql(query, tuple(params))]

    if not matching_contact_names:
        return {"count": 0, "headers": ["Nombre"], "contacts": []}

    contact_filters = {
        "name": ["in", matching_contact_names], 
        "status": ["in", ["Enabled", "Passive"]],
        "custom_is_liseniq_contact": 1,
        "custom_is_deleted": 0,
        "custom_status": ["!=", "Inactivo"]
    }
    if user_contact_name:
        contact_filters["name"] = ["in", [name for name in matching_contact_names if name != user_contact_name]]

    if user_company:
        contact_filters["custom_company"] = user_company


    contact_docs = frappe.get_all(
        "Contact",
        filters=contact_filters,
        fields=["name", "first_name", "last_name"],
        ignore_permissions=True
    )
    
    demographic_map_docs = frappe.get_all(
        "qp_IQ_DemographicType",
        filters={"name": ["in", list(set(demographic_ids))]},
        fields=["name", "dt_title"]
    )
    id_to_title_map = {doc.name: doc.dt_title for doc in demographic_map_docs}
    
    headers = ["Nombre"] + [id_to_title_map[demo_id] for demo_id in demographic_ids if demo_id in id_to_title_map]

    demographic_details = frappe.get_all(
        "qp_IQ_ContactAdditionalDetail",
        filters={"parent": ["in", [c.name for c in contact_docs]]},
        fields=["parent", "cad_demographic_type", "cad_value"]
    )

    details_map = {}
    for detail in demographic_details:
        if detail.parent not in details_map:
            details_map[detail.parent] = {}
        if detail.cad_demographic_type in id_to_title_map:
            demographic_title = id_to_title_map[detail.cad_demographic_type]
            details_map[detail.parent][demographic_title] = detail.cad_value

    contacts_for_modal = []
    for contact in contact_docs:
        contact_data = {"name": contact.name, "Nombre": f"{contact.first_name} {contact.last_name or ''}".strip()}
        contact_specific_details = details_map.get(contact.name, {})
        for header in headers:
            if header != "Nombre":
                contact_data[header] = contact_specific_details.get(header, "N/A")
        contacts_for_modal.append(contact_data)

    return {"count": len(contact_docs), "headers": headers, "contacts": contacts_for_modal}

@frappe.whitelist()
def delete_measurement_contacts(survey_name, contact_names):
    try:
        if isinstance(contact_names, str):
            try:
                contact_names = json.loads(contact_names)
            except Exception:
                contact_names = [contact_names]
        if not isinstance(contact_names, (list, tuple)) or not contact_names:
            return {"status": "error", "message": _("Lista de contactos inválida.")}

        survey = frappe.get_doc("qp_IQ_Survey", survey_name)
        user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")
        if not user_company or survey.su_owner != user_company:
            return {"status": "error", "message": _("No tiene permisos para modificar esta medición.")}

        removed_no_response = []
        removed_with_response = []
        not_found = []
        errors = []

        for contact_name in contact_names:
            try:
                rec = frappe.db.get_value(
                    "qp_IQ_SurveyRecipient",
                    {"sr_survey": survey.name, "sr_contact": contact_name},
                    ["name", "sr_status"],
                    as_dict=True
                )

                if not rec:
                    not_found.append(contact_name)
                    continue

                responded = (rec.get("sr_status") == "Responded")
                if responded:
                    found_responses = frappe.get_all(
                        "Survey Response",
                        filters={"survey": survey.su_name, "user": contact_name},
                        fields=["name"]
                    )
                    for r in found_responses:
                        try:
                            frappe.delete_doc("Survey Response", r.name, ignore_permissions=True)
                        except Exception as e_del_resp:
                            errors.append(f"Contacto {contact_name}: no se pudo eliminar Survey Response {r.name}: {e_del_resp}")

                    frappe.delete_doc("qp_IQ_SurveyRecipient", rec.name, ignore_permissions=True)
                    removed_with_response.append(contact_name)
                else:
                    frappe.delete_doc("qp_IQ_SurveyRecipient", rec.name, ignore_permissions=True)
                    removed_no_response.append(contact_name)

            except Exception as e_item:
                errors.append(f"{contact_name}: {e_item}")

        frappe.db.commit()
        return {
            "status": "success",
            "removed_without_response": removed_no_response,
            "removed_with_response": removed_with_response,
            "not_found": not_found,
            "errors": errors
        }

    except Exception as e:
        frappe.db.rollback()
        frappe.log_error(frappe.get_traceback(), "delete_measurement_contacts")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def save_measurement(data):
    try:
        data = json.loads(data)
        email_data = data.get("email_customization") or {}
        email_use_default = bool(data.get("email_use_default"))
        
        final_survey_name = None

        if data.get("is_edit_mode") and data.get("doc_name"):
            survey = frappe.get_doc("qp_IQ_Survey", data["doc_name"])

            new_name = data.get("name")
            if new_name:
                exists = frappe.get_all(
                    "qp_IQ_Survey",
                    filters=[
                        ["qp_IQ_Survey", "su_owner", "=", survey.su_owner],
                        ["qp_IQ_Survey", "su_name", "=", new_name],
                        ["qp_IQ_Survey", "name", "!=", survey.name],
                    ],
                    fields=["name"],
                    limit_page_length=1
                )
                if exists:
                    return {"status": "error", "message": _("Ya existe una medición con ese nombre para su empresa.")}

                survey.su_name = new_name

            survey.su_end_date = data.get("endDate")
            if data.get("startDate"):
                survey.su_start_date = data.get("startDate")

            reminders = data.get("reminders")
            if reminders:
                survey.su_send_reminders = 1
                survey.su_reminder_frequency = reminders.get("frequency")
                survey.su_reminder_max = reminders.get("max")
            else:
                survey.su_send_reminders = 0
                survey.su_reminder_frequency = None
                survey.su_reminder_max = None

            survey.su_is_leadership = 1 if data.get("is_leadership") else 0
            survey.su_invitation_subject = email_data.get("invitation_subject")
            survey.su_invitation_body = email_data.get("invitation_body")
            survey.su_reminder_subject = email_data.get("reminder_subject")
            survey.su_reminder_body = email_data.get("reminder_body")
            survey.su_default_notif = "1" if email_use_default else "0"

            survey.save(ignore_permissions=True)
            frappe.db.commit()
            final_survey_name = survey.name

            # Lógica de actualización de Contactos
            rs_not_sent = frappe.get_value("qp_IQ_RecipientStatus", {"rs_status": "Not Sent"}, "name") or "Not Sent"
            
            if survey.su_is_leadership:
                leadership_network = data.get("leadershipNetwork", [])
                if leadership_network:
                    existing_recipients = frappe.get_all("qp_IQ_SurveyRecipient", filters={"sr_survey": survey.name}, fields=["name", "sr_contact", "sr_evaluating_to", "sr_evaluation_role"])
                    existing_pairs = {(r.sr_evaluating_to, r.sr_contact, r.sr_evaluation_role) for r in existing_recipients}
                    
                    for network in leadership_network:
                        leader_id = network.get("leader", {}).get("id")
                        for ev in network.get("evaluators", []):
                            evaluator_id = ev.get("id")
                            role_label = ev.get("role")
                            
                            if (evaluator_id, leader_id, role_label) not in existing_pairs:
                                has_responded = frappe.db.exists("Survey Response", {"survey": survey.su_name, "user": evaluator_id, "custom_evaluatee": leader_id})
                                if not has_responded:
                                    doc_vals = {
                                        "doctype": "qp_IQ_SurveyRecipient",
                                        "sr_survey": survey.name,
                                        "sr_contact": leader_id,
                                        "sr_evaluating_to": evaluator_id,
                                        "sr_status": rs_not_sent
                                    }
                                    if not ev.get("isAuto") and role_label:
                                        doc_vals["sr_evaluation_role"] = role_label
                                    elif ev.get("isAuto"):
                                        doc_vals["sr_evaluation_role"] = "Autoevaluación"
                                    
                                    frappe.get_doc(doc_vals).insert(ignore_permissions=True)

            elif data.get("contacts", {}).get("surveyType") != "anonymous_link":
                contacts_data = data.get("contacts", {})
                new_contacts = contacts_data.get("list", [])
                if new_contacts:
                    existing_recipients = frappe.get_all(
                        "qp_IQ_SurveyRecipient",
                        filters={"sr_survey": survey.name},
                        fields=["sr_contact"]
                    )
                    existing_contact_names = set(r["sr_contact"] for r in existing_recipients if r["sr_contact"])

                    for contact in new_contacts:
                        contact_name = contact.get("name")
                        if not contact_name or contact_name in existing_contact_names:
                            continue

                        has_responded = frappe.db.exists(
                            "Survey Response",
                            {"survey": survey.su_name, "user": contact_name}
                        )
                        if has_responded:
                            continue

                        frappe.get_doc({
                            "doctype": "qp_IQ_SurveyRecipient",
                            "sr_survey": survey.name,
                            "sr_contact": contact_name,
                            "sr_status": rs_not_sent
                        }).insert(ignore_permissions=True)
                        
            frappe.db.commit()

            try:
                frappe.call("liseniq.tasks.send_pending_links_for_survey", survey_name=survey.name)
            except Exception:
                frappe.log_error(frappe.get_traceback(), "save_measurement.send_pending_links_on_edit")

        else:
            question_types_map = {qt.name: qt for qt in frappe.get_all("qp_IQ_QuestionType", fields=["name", "qnt_type_name", "qnt_mnemonico"])}
            user_contact = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "name")
            user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")

            manual_question_map = {}
            if data.get("questions"):
                for q in data["questions"]:
                    if q.get("id", "").startswith("manual-"):
                        question_text = q["text"]
                        existing_question = frappe.db.exists("qp_IQ_Question", {"qn_statement": question_text, "qn_owner": user_company})

                        if existing_question:
                            manual_question_map[q["id"]] = existing_question
                        else:
                            new_question = frappe.new_doc("qp_IQ_Question")
                            new_question.qn_statement = question_text
                            new_question.qn_type = q["type"]
                            new_question.qn_creator = user_contact
                            new_question.qn_owner = user_company
                            
                            if q.get("demographic"):
                                demographic_name = frappe.db.exists("qp_IQ_DemographicType", {"dt_title": q["demographic"]})
                                if not demographic_name:
                                    demographic_doc = frappe.new_doc("qp_IQ_DemographicType")
                                    demographic_doc.dt_title = q["demographic"]
                                    demographic_doc.dt_object_type = "Pregunta"
                                    demographic_doc.insert(ignore_permissions=True)
                                    demographic_name = demographic_doc.name
                                new_question.qn_demographic = demographic_name

                            if q.get("options"):
                                if q.get("typeName") == "Likert" or (isinstance(q.get("options")[0], dict) and "value" in q["options"][0]):
                                    for opt in q["options"]:
                                        new_question.append("qn_response_options", {
                                            "qo_option_text": opt["text"] if isinstance(opt, dict) else opt,
                                            "qo_option_value": opt["value"] if isinstance(opt, dict) and "value" in opt else opt,
                                            "qo_url": (opt.get("url") if isinstance(opt, dict) and "url" in opt else None)
                                        })
                                else:
                                    for opt_text in q["options"]:
                                        new_question.append("qn_response_options", {"qo_option_text": opt_text, "qo_option_value": opt_text})
                            
                            if q.get("negative_statement"): new_question.qn_statement_negative = q["negative_statement"]
                            if q.get("positive_statement"): new_question.qn_statement_positive = q["positive_statement"]
                            if q.get("nps_min") is not None: new_question.qn_nps_min = q["nps_min"]
                            if q.get("nps_max") is not None: new_question.qn_nps_max = q["nps_max"]

                            if q.get("qp_others"): new_question.qp_others = 1
                            if q.get("qp_none_above"): new_question.qp_none_above = 1
                                
                            new_question.insert(ignore_permissions=True)
                            manual_question_map[q["id"]] = new_question.name

            surveyjs_doc_name = None
            if data.get("questions"):
                elements = []

                LIKERT_ICON_MAP = {
                    5: "/files/aiq - totalmente de acuerdo.png",
                    4: "/files/aiq - de acuerdo.png",
                    3: "/files/aiq - ni de acuerdo ni desacuerdo.png",
                    2: "/files/aiq - desacuerdo.png",
                    1: "/files/aiq - totalmente desacuerdo.png",
                }
                for q in data["questions"]:
                    question_name = manual_question_map.get(q["id"]) if q.get("id", "").startswith("manual-") else q["id"]
                    qt_info = question_types_map.get(q["type"], {})
                    question_type_mnemonic = qt_info.get("qnt_mnemonico")
                    
                    surveyjs_type = "text"
                    if question_type_mnemonic == "radio_group": surveyjs_type = "radiogroup"
                    elif question_type_mnemonic == "text_area": surveyjs_type = "comment"
                    elif question_type_mnemonic == "score_nps": surveyjs_type = "rating"
                    elif question_type_mnemonic in ["scale_likert", "scale_emoji"]: surveyjs_type = "imagepicker"
                    elif question_type_mnemonic == "check_group": surveyjs_type = "checkbox"
                    elif question_type_mnemonic == "text_short": surveyjs_type = "text"

                    element = {
                        "type": surveyjs_type,
                        "name": question_name,
                        "title": q["text"],
                        "isRequired": "true"
                    }

                    if question_type_mnemonic == "scale_likert":
                        choices = []
                        try:
                            if question_name:
                                q_doc = frappe.get_doc("qp_IQ_Question", question_name)
                                if q_doc and q_doc.qn_response_options:
                                    for opt in q_doc.qn_response_options:
                                        try: val_int = int(opt.qo_option_value)
                                        except Exception: val_int = None
                                        image_url = getattr(opt, "qo_url", None) or (LIKERT_ICON_MAP.get(val_int) if val_int else None)
                                        choices.append({
                                            "text": opt.qo_option_text,
                                            "value": opt.qo_option_value,
                                            "imageLink": image_url if image_url else "data:image/gif;base64,R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7"
                                        })
                        except Exception:
                            choices = []
                        element["choices"] = choices
                        element["showLabel"] = True
                        element["multiSelect"] = False
                        element["imageFit"] = "contain"
                        element["imageHeight"] = 32
                        element["imageWidth"] = 32
                        element["choicesOrder"] = "none"

                    elif question_type_mnemonic == "scale_emoji":
                        choices = []
                        try:
                            if question_name:
                                q_doc = frappe.get_doc("qp_IQ_Question", question_name)
                                if q_doc and q_doc.qn_response_options:
                                    for opt in q_doc.qn_response_options:
                                        choices.append({
                                            "text": opt.qo_option_text,
                                            "value": opt.qo_option_value,
                                            "imageLink": opt.qo_url if opt.qo_url else "data:image/gif;base64,R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7"
                                        })
                        except Exception:
                            choices = []
                        element["choices"] = choices
                        element["showLabel"] = True
                        element["multiSelect"] = False
                        element["imageFit"] = "contain"
                        element["imageHeight"] = 32
                        element["imageWidth"] = 32
                        element["choicesOrder"] = "none"

                    elif question_type_mnemonic == "radio_group" and q.get("options"):
                        element["choices"] = q["options"]
                        
                    elif question_type_mnemonic == "check_group":
                        if q.get("options"):
                            element["choices"] = q["options"]
                        if q.get("qp_others"):
                            element["hasOther"] = True
                            element["otherText"] = "Otros"
                        if q.get("qp_none_above"):
                            element["hasNone"] = True
                            element["noneText"] = "Ninguna de las anteriores"

                    elif question_type_mnemonic == "score_nps":
                        element["rateMin"] = q.get("nps_min", 1)
                        element["rateMax"] = q.get("nps_max", 10)
                        element["minRateDescription"] = "NADA PROBABLE"
                        element["maxRateDescription"] = "MUY PROBABLE"
                        element["rateDescriptionLocation"] = "top"
                        
                    elif question_type_mnemonic == "text_short":
                        element["maxLength"] = 70

                    elements.append(element)

                survey_json_content = {
                    "title": data["name"],
                    "description": "",
                    "pages": [{"name": "page1", "elements": elements}]
                }

                surveyjs_doc = frappe.new_doc("Survey")
                surveyjs_doc.name = data["name"]
                surveyjs_doc.title = data["name"]
                surveyjs_doc.survey_json = json.dumps(survey_json_content)
                surveyjs_doc.insert(ignore_permissions=True)
                surveyjs_doc_name = surveyjs_doc.name

                web_form_route = data["name"].lower().replace(" ", "-")           
                web_form = frappe.new_doc("Web Form")
                web_form.title = data["name"]
                web_form.route = web_form_route
                web_form.doc_type = "Survey Response"
                web_form.module = "Frappe Survey"
                web_form.client_script = WEB_FORM_CLIENT_SCRIPT
                web_form.custom_css = WEB_FORM_CUSTOM_CSS
                web_form.published = 1

                survey_response_meta = frappe.get_meta("Survey Response")
                fieldtype_mapping = {"Long Text": "Text Editor"}
                for field in survey_response_meta.fields:
                    if field.fieldtype not in ["Section Break", "Column Break", "Tab Break"]:
                        web_form_fieldtype = fieldtype_mapping.get(field.fieldtype, field.fieldtype)
                        
                        is_hidden = field.hidden
                        if field.fieldname in ["custom_evaluatee", "custom_evaluator", "responses", "response_json", "user", "survey"]:
                            is_hidden = 1

                        web_form.append("web_form_fields", {
                            "fieldname": field.fieldname,
                            "fieldtype": web_form_fieldtype,
                            "label": field.label,
                            "reqd": field.reqd,
                            "options": field.options,
                            "hidden": is_hidden,
                            "read_only": field.read_only,
                            "default": field.default,
                            "description": field.description,
                        })
                web_form.insert(ignore_permissions=True)

            user_contact_info = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")
            if not user_contact_info:
                message = "El usuario actual no tiene una compañía asignada para definir la propiedad de la medición."
                frappe.log_error(message, "Error en save_measurement")
                return {"status": "error", "message": message}
            user_company = user_contact_info
            
            default_status_text = "Programada"
            status_name = frappe.db.get_value("qp_IQ_SurveyStatus", {"se_status": default_status_text}, "name")
            
            if not status_name:
                status_doc = frappe.new_doc("qp_IQ_SurveyStatus")
                status_doc.se_status = default_status_text
                status_doc.insert(ignore_permissions=True)
                status_name = status_doc.name
            
            survey = frappe.new_doc("qp_IQ_Survey")
            survey.su_name = data["name"]
            survey.su_owner = user_company
            survey.su_is_leadership = 1 if data.get("is_leadership") else 0
            survey.su_start_date = data.get("startDate")
            survey.su_end_date = data.get("endDate")
            survey.su_timezone = data.get("timezone")
            
            template_cache_key = f"measurement_template:{frappe.session.user}"
            template_name = (
                data.get("template_name")
                or frappe.request.args.get("template")
                or getattr(frappe.local, "template_name", None)
                or frappe.cache().get_value(template_cache_key)
            )
            if template_name:
                survey.su_template = template_name
                frappe.cache().delete_value(template_cache_key)
            
            contacts_data = data.get("contacts", {})
            survey_type = contacts_data.get("surveyType")
            response_type = contacts_data.get("responseType")

            if survey_type == 'anonymous_link':
                survey.su_is_anonymous = 1
                survey.su_custom_generate_public_link = 1
            else:
                survey.su_is_anonymous = 0 
                if survey_type == 'all':
                    survey.su_custom_generate_public_link = 0
                elif survey_type == 'selected':
                    survey.su_custom_generate_public_link = 1

            survey.su_status = status_name
            if data.get("reminders"):
                survey.su_send_reminders = 1
                survey.su_reminder_frequency = data["reminders"]["frequency"]
                survey.su_reminder_max = data["reminders"]["max"]
            else:
                survey.su_send_reminders = 0

            survey.su_invitation_subject = email_data.get("invitation_subject")
            survey.su_invitation_body = email_data.get("invitation_body")
            survey.su_reminder_subject = email_data.get("reminder_subject")
            survey.su_reminder_body = email_data.get("reminder_body")
            survey.su_default_notif = "1" if email_use_default else "0"

            if data.get("questions"):
                for q in data["questions"]:
                    question_name = manual_question_map.get(q["id"]) if q.get("id", "").startswith("manual-") else q["id"]
                    if question_name:
                        survey.append("su_questions", {"sq_question": question_name})

            survey.insert(ignore_permissions=True)
            frappe.db.commit()

            final_survey_name = survey.name
            rs_not_sent = frappe.get_value("qp_IQ_RecipientStatus", {"rs_status": "Not Sent"}, "name") or "Not Sent"

            if survey.su_is_leadership:
                leadership_network = data.get("leadershipNetwork", [])
                for network in leadership_network:
                    leader_id = network.get("leader", {}).get("id")
                    for ev in network.get("evaluators", []):
                        evaluator_id = ev.get("id")
                        role_label = ev.get("role")
                        
                        doc_vals = {
                            "doctype": "qp_IQ_SurveyRecipient",
                            "sr_survey": survey.name,
                            "sr_contact": leader_id,
                            "sr_evaluating_to": evaluator_id,
                            "sr_status": rs_not_sent
                        }
                        if not ev.get("isAuto") and role_label:
                            doc_vals["sr_evaluation_role"] = role_label
                        elif ev.get("isAuto"):
                            doc_vals["sr_evaluation_role"] = "Autoevaluación"
                            
                        frappe.get_doc(doc_vals).insert(ignore_permissions=True)
                frappe.db.commit()

            else:
                if survey_type == 'selected' and contacts_data.get("list"):
                    contact_names = [c.get("name") for c in contacts_data.get("list") if c.get("name")]
                    if contact_names:
                        for contact_name in contact_names:
                            frappe.get_doc({
                                "doctype": "qp_IQ_SurveyRecipient",
                                "sr_survey": survey.name,
                                "sr_contact": contact_name,
                                "sr_status": rs_not_sent
                            }).insert(ignore_permissions=True)
                    frappe.db.commit()

        frappe.db.commit()
        return {"status": "success", "message": f"Medición '{survey.su_name}' creada exitosamente.", "docname": survey.name}
    
    except Exception as e:
        frappe.db.rollback()
        frappe.log_error(frappe.get_traceback(), "Error en save_measurement")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def generate_leadership_excel_template():
    """ Genera la plantilla Excel para carga masiva de red 360° """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Carga_Masiva_360"
        
        # Headers principales
        ws.append(["email", "relacion"])
        
        # Datos de ejemplo
        ws.append(["ejemplo_lider@empresa.com", "Autoevaluación"])
        ws.append(["ejemplo_evaluador@empresa.com", "Evaluador"])
        
        # Recuperar Roles activos de Catalog Options
        catalog = frappe.db.get_value("qp_IQ_Catalog", {"ca_mnemonico": "measurement_roles"}, "name")
        roles = []
        if catalog:
            roles_data = frappe.get_all("qp_IQ_CatalogOptions", filters={"co_catalog": catalog, "co_is_active": 1}, fields=["co_label"])
            roles = [r.co_label for r in roles_data]
            
        # Segunda hoja con instrucciones y roles válidos
        ws_info = wb.create_sheet(title="Roles_Validos")
        ws_info.append(["Información Importante:"])
        ws_info.append(["- Inicie el bloque con el líder usando 'Autoevaluación' en la columna relacion."])
        ws_info.append(["- Las siguientes filas serán los evaluadores asignados a ese líder."])
        ws_info.append([])
        ws_info.append(["ROLES PERMITIDOS (copiar exactamente):"])
        ws_info.append(["Autoevaluación"])
        
        row_idx = 7
        for r in roles:
            ws_info.append([r])
            row_idx += 1

        # Añadir lista desplegable en la columna relacion (B) de la hoja principal
        formula = f"='Roles_Validos'!$A$6:$A${row_idx - 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = 'Debe seleccionar un rol válido de la lista desplegable.'
        dv.errorTitle = 'Rol Inválido'
        dv.prompt = 'Seleccione un rol de la lista.'
        dv.promptTitle = 'Relación'
        
        ws.add_data_validation(dv)
        dv.add('B2:B1000')

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return base64.b64encode(stream.read()).decode('utf-8')
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "generate_leadership_excel_template")
        return None

@frappe.whitelist()
def process_leadership_excel(file_base64):
    """ Procesa el Excel subido en base64 para construir la red 360 """
    try:
        # Decodificar el archivo Excel
        file_data = base64.b64decode(file_base64.split(",")[1])
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return {"status": "error", "message": "El archivo está vacío o no contiene datos válidos."}

        headers = [str(h).lower().strip() for h in rows[0] if h]
        if "email" not in headers or "relacion" not in headers:
            return {"status": "error", "message": "El archivo Excel debe contener exactamente las columnas 'email' y 'relacion'."}
        
        email_idx = headers.index("email")
        relacion_idx = headers.index("relacion")

        # Obtener todos los contactos de la empresa del usuario actual
        user_company = frappe.db.get_value("Contact", {"user": frappe.session.user, "custom_is_liseniq_contact": 0}, "custom_company")
        contacts = frappe.db.sql("""
            SELECT c.name, CONCAT(IFNULL(c.first_name, ''), ' ', IFNULL(c.last_name, '')) as full_name, ce.email_id
            FROM `tabContact` c
            LEFT JOIN `tabContact Email` ce ON ce.parent = c.name AND ce.parenttype = 'Contact'
            WHERE c.status IN ('Enabled', 'Passive')
            AND c.custom_is_liseniq_contact = 1
            AND c.custom_is_deleted = 0
            AND c.custom_company = %s
        """, (user_company,), as_dict=True)

        # Mapeo de emails (minúsculas) a la data de contacto
        email_map = {c.email_id.lower().strip(): {"id": c.name, "name_display": c.full_name.strip()} for c in contacts if c.email_id}

        # Roles válidos y activos
        catalog = frappe.db.get_value("qp_IQ_Catalog", {"ca_mnemonico": "measurement_roles"}, "name")
        valid_roles = []
        if catalog:
            roles_data = frappe.get_all("qp_IQ_CatalogOptions", filters={"co_catalog": catalog, "co_is_active": 1}, fields=["co_label"])
            valid_roles = [r.co_label for r in roles_data]
            
        valid_roles_lower = {r.lower(): r for r in valid_roles}

        networks = []
        current_leader = None
        evaluators = []
        errors = []

        # Recorrer las filas del Excel
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(email_idx, relacion_idx):
                continue
                
            email = str(row[email_idx]).strip().lower() if row[email_idx] else None
            relacion = str(row[relacion_idx]).strip() if row[relacion_idx] else None

            if not email and not relacion:
                continue
                
            if not email:
                errors.append(f"Fila {i}: Falta el correo electrónico.")
                continue
                
            if not relacion:
                errors.append(f"Fila {i}: Falta la relación para {email}.")
                continue

            contact_info = email_map.get(email)
            if not contact_info:
                errors.append(f"Fila {i}: El correo '{email}' no corresponde a un contacto activo en su empresa.")
                continue

            relacion_lower = relacion.lower()

            if relacion_lower in ["autoevaluación", "autoevaluacion"]:
                # Guardar el bloque del líder anterior si existe
                if current_leader:
                    networks.append({"leader": current_leader, "evaluators": evaluators})
                
                # Iniciar un nuevo bloque de liderazgo
                current_leader = contact_info
                evaluators = [{
                    "id": contact_info["id"],
                    "name": contact_info["name_display"],
                    "role": "Autoevaluación",
                    "role_id": "Autoevaluación",
                    "isAuto": True
                }]
            else:
                if not current_leader:
                    errors.append(f"Fila {i}: Evaluador '{email}' encontrado antes de definir un líder (Se requiere una fila de 'Autoevaluación' primero).")
                    continue
                    
                if relacion_lower in valid_roles_lower:
                    real_role = valid_roles_lower[relacion_lower]
                    
                    # Evitar duplicados del mismo evaluador para un mismo líder
                    if not any(e["id"] == contact_info["id"] for e in evaluators):
                        evaluators.append({
                            "id": contact_info["id"],
                            "name": contact_info["name_display"],
                            "role": real_role,
                            "role_id": real_role,
                            "isAuto": False
                        })
                else:
                    errors.append(f"Fila {i}: La relación '{relacion}' no es válida o está inactiva.")

        # Guardar el último bloque leído
        if current_leader:
            networks.append({"leader": current_leader, "evaluators": evaluators})

        return {
            "status": "success", 
            "networks": networks,
            "errors": errors
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "process_leadership_excel")
        return {"status": "error", "message": str(e)}