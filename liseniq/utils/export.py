import os

import frappe
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from frappe.utils.pdf import get_pdf
from openpyxl import Workbook
from openpyxl.styles import Alignment
from frappe.utils.background_jobs import enqueue
from frappe.utils.file_manager import save_file
from rq.job import Job
from redis import Redis

import re
from bs4 import BeautifulSoup



def _sanitize_filename(value):
    value = (value or "").strip()
    safe_chars = []
    for char in value:
        if char.isalnum() or char in (" ", "-", "_", "."):
            safe_chars.append(char)
        else:
            safe_chars.append("_")
    sanitized = "".join(safe_chars).strip().replace(" ", "_")
    return sanitized or "reporte"


def _get_survey_doc(survey_identifier):
    try:
        return frappe.get_doc("qp_IQ_Survey", survey_identifier)
    except Exception:
        return frappe.get_doc("qp_IQ_Survey", {"su_name": survey_identifier})


def _leadership_pdf_options():
    return {
        "page-size": "A4",
        "margin-top": "0mm",
        "margin-bottom": "0mm",
        "margin-left": "0mm",
        "margin-right": "0mm",
        "header-spacing": "0",
        "disable-smart-shrinking": "",
        "enable-local-file-access": "",  
        "dpi": "300",  
        "print-media-type": ""
    }


def _ensure_pdf_header_footer_placeholders(html):
    if "id=\"header-html\"" in html and "id=\"footer-html\"" in html:
        return html

    placeholders = "<div id=\"header-html\"></div><div id=\"footer-html\"></div>"
    if "</body>" in html:
        return html.replace("</body>", placeholders + "</body>", 1)
    return html + placeholders


def _compile_css_for_pdf(html):



    css_variables = {
        "--brand-primary": "#502394",
        "--brand-secondary": "#14B8A6",
        "--brand-primary-light": "#f3e8ff",
        "--brand-light-gray": "#FAF9F8",
        "--brand-border-color": "#dee2e6",
        "--text-color-primary": "#212529",
        "--text-color-secondary": "#6c757d",
    }
    
    for var_name, var_value in css_variables.items():
        html = html.replace(f"var({var_name})", var_value)
    
    style_pattern = r'<style[^>]*>(.*?)</style>'
    
    def process_style_tag(match):
        style_content = match.group(1)
        

        before_pattern = r'([\w\s\-\.,#:>]+)::before\s*\{([^}]*)\}'
        style_content = re.sub(before_pattern, lambda m: _handle_pseudo_element(m, "before"), style_content)

        after_pattern = r'([\w\s\-\.,#:>]+)::after\s*\{([^}]*)\}'
        style_content = re.sub(after_pattern, lambda m: _handle_pseudo_element(m, "after"), style_content)
        
        return f'<style>{style_content}</style>'
    
    html = re.sub(style_pattern, process_style_tag, html, flags=re.DOTALL)

    z_pattern = r'z-index:\s*(\d+)'
    
    def limit_z_index(match):
        z_value = int(match.group(1))
        limited_z = min(z_value, 999)
        return f'z-index: {limited_z}'
    
    html = re.sub(z_pattern, limit_z_index, html)
    
    return html


def _handle_pseudo_element(match, pseudo_type):
    """Convierte pseudoelementos a estilos inline """
    selector = match.group(1).strip()
    properties = match.group(2).strip()
    
    return f'{selector} {{ {properties} }}'

@frappe.whitelist()
def export_survey_results(survey_name):
    current_user = frappe.session.user
    frappe.session.user = "Administrator"
    try:
        frappe.flags.ignore_permissions = True
        survey = frappe.get_doc("qp_IQ_Survey", survey_name)
        survey_name = survey.su_name
        is_leadership = survey.su_is_leadership
        if is_leadership:
            REPORT_NAME = "Engagement Responses"
        else:
            REPORT_NAME = "Survey Response Custom Report Front"
        filters = {"survey": survey_name}

        report = frappe.get_doc("Report", REPORT_NAME)
        columns, data = report.get_data(filters)

        # Separar preguntas abiertas sin modificar el array original mientras se itera
        data_open_questions = []
        data_closed = []

        for row in data:
            variable = row.get('variable', '').lower()
            if "abierta" in variable or "abiertas" in variable:
                data_open_questions.append(row)
            else:
                data_closed.append(row)

        # Extraer headers de columnas
        col_labels = [col.get('label', col.get('fieldname', '')) for col in columns]
        col_fields = [col.get('fieldname', '') for col in columns]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resultados de la encuesta"
        sheet.append(col_labels)

        for row in data_closed:
            sheet.append([row.get(field, '') for field in col_fields])

        if data_open_questions:
            open_sheet = workbook.create_sheet(title="Preguntas Abiertas")
            open_sheet.append(col_labels)
            for row in data_open_questions:
                open_sheet.append([row.get(field, '') for field in col_fields])

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        filename = f"{survey_name}_results.xlsx"
        
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = excel_file.read()
        frappe.local.response.type = "download"
        frappe.local.response.display_content_as = "attachment"
    finally:
        frappe.flags.ignore_permissions = False
        frappe.session.user = current_user

@frappe.whitelist()
def get_demographics(survey):
    """Obtiene la lista de demográficos de tipo Contacto"""
    survey_owner = frappe.get_value("qp_IQ_Survey", {"name": survey}, "su_owner")
    demographics = frappe.get_all(
        'qp_IQ_DemographicType',
        filters={'dt_object_type': 'Contacto', 'dt_creator_company': survey_owner},
        fields=['name', 'dt_title'],
        order_by='dt_title'
    )
    return demographics

@frappe.whitelist()
def export_follow_up_report(survey_name, demographic1=None, demographic2=None):
    current_user = frappe.session.user
    frappe.session.user = "Administrator"
    try:
        frappe.flags.ignore_permissions = True
        survey = frappe.get_doc("qp_IQ_Survey", survey_name)
        survey_id = survey.name
        survey_display_name = survey.su_name
        
        REPORT_NAME = "Survey Status"
        filters = {"survey": survey_id}

        if demographic1:
            filters["demographic1"] = demographic1
        if demographic2:
            filters["demographic2"] = demographic2

        report = frappe.get_doc("Report", REPORT_NAME)
        columns, data = report.get_data(filters)

        col_labels = [col.get('label', col.get('fieldname', '')) for col in columns]
        col_fields = [col.get('fieldname', '') for col in columns]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Seguimiento"
        sheet.append(col_labels)

        for row in data:
            sheet.append([row.get(field, '') for field in col_fields])

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        filename = f"{survey_display_name}_seguimiento.xlsx"
        
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = excel_file.read()
        frappe.local.response.type = "download"
        frappe.local.response.display_content_as = "attachment"
    finally:
        frappe.flags.ignore_permissions = False
        frappe.session.user = current_user


@frappe.whitelist()
def export_seguimiento_360(survey_name):
    current_user = frappe.session.user
    frappe.session.user = "Administrator"
    try:
        frappe.flags.ignore_permissions = True
        survey = frappe.get_doc("qp_IQ_Survey", {"su_name": survey_name})
        survey_id = survey.name
        survey_display_name = survey.su_name
        
        REPORT_NAME = "Seguimiento Mediciones 360"
        filters = {"survey": survey_name}

        report = frappe.get_doc("Report", REPORT_NAME)
        columns, data = report.get_data(filters)

        col_labels = [col.get('label', col.get('fieldname', '')) for col in columns]
        col_fields = [col.get('fieldname', '') for col in columns]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Seguimiento 360"
        sheet.append(col_labels)

        for row in data:
            sheet.append([row.get(field, '') for field in col_fields])

        # Combinar celdas para cada evaluado
        current_evaluated = None
        start_row = 2  # Empezamos en la fila 2 porque la fila 1 tiene los encabezados
        for idx, row in enumerate(data, start=2):
            evaluated = row.get('evaluated')
            if evaluated != current_evaluated:
                if current_evaluated is not None:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=idx-1, end_column=1)
                    sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                current_evaluated = evaluated
                start_row = idx
        # Merge para el último evaluado
        if current_evaluated is not None:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=idx, end_column=1)
            sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        filename = f"{survey_display_name}_seguimiento_360.xlsx"
        
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = excel_file.read()
        frappe.local.response.type = "download"
        frappe.local.response.display_content_as = "attachment"
    finally:
        frappe.flags.ignore_permissions = False
        frappe.session.user = current_user


@frappe.whitelist()
def export_leadership_reports_zip(survey_name):
    frappe.cache().delete_value(f"export_job_{survey_name}")

    job = enqueue(
        _generate_zip_job,
        queue="long",
        timeout=1800,
        survey_name=survey_name,
        enqueued_by=frappe.session.user,
    )
    return {"job_id": job.id, "cache_key": f"export_job_{survey_name}"}



@frappe.whitelist()
def get_export_job_status(job_id, cache_key):
    cached = frappe.cache().get_value(cache_key)
    if cached:
        return cached

    try:

        redis_conn = Redis.from_url(frappe.conf.redis_queue)
        job = Job.fetch(job_id, connection=redis_conn)
        status = job.get_status()

        if status == "failed":
            return {"status": "failed", "error": str(job.exc_info)}

        return {"status": str(status)}

    except Exception as e:
        return {"status": "error", "error": str(e)}



def _generate_zip_job(survey_name, enqueued_by):
    frappe.set_user("Administrator")
    frappe.flags.ignore_permissions = True

    try:
        survey = _get_survey_doc(survey_name)
        survey_display_name = survey.su_name
        template_category_id = frappe.db.get_value("qp_IQ_Template", {"name": survey.su_template}, "tp_category")
        template_category = frappe.db.get_value("qp_IQ_TemplateCategory", {"name": template_category_id}, "qnc_category")
        if not survey.su_is_leadership:
            if template_category == "Engagement":
                doctype_name   = "qp_IQ_Cultura_Report"
                filter_field   = "cutoff_name"
                print_format   = "Reporte de Engagement"
            else:
                doctype_name   = "qp_IQ_Cultura_Report"
                filter_field   = "cutoff_name"
                print_format   = "Reporte de Cultura"
        else:
            doctype_name   = "qp_IQ_Leader_360_Report"
            filter_field   = "leader_name"
            print_format   = "Reporte Individual Liderazgo"

        reports = frappe.get_all(
            doctype_name,
            filters={"survey_name": survey_display_name},
            fields=["name", filter_field],
            order_by=filter_field + " asc",
        )

        if not reports:
            frappe.throw("No se encontraron informes para esta medición.")

        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, "w", ZIP_DEFLATED) as zip_file:
            for report_row in reports:
                pdf_bytes = _render_report_pdf(doctype_name, report_row, print_format)
                file_name = _resolve_file_name(survey, doctype_name, report_row)
                zip_file.writestr(f"{file_name}.pdf", pdf_bytes)

        zip_buffer.seek(0)
        filename = f"{_sanitize_filename(survey_display_name)}_informes.zip"

        saved = save_file(
            fname=filename,
            content=zip_buffer.read(),
            dt='qp_IQ_Survey',
            dn=survey_name,
            is_private=1,
        )
        frappe.db.commit()

        file_url = saved.file_url
        frappe.cache().set_value(
            f"export_job_{survey_name}",
            {"status": "finished", "file_url": file_url},
            expires_in_sec=3600,
        )

        return file_url

    except Exception as e:
        frappe.cache().set_value(
            f"export_job_{survey_name}",
            {"status": "failed", "error": str(e)},
            expires_in_sec=3600,
        )
        raise

    finally:
        frappe.flags.ignore_permissions = False

@frappe.whitelist()
def download_export_file(cache_key):
    cached = frappe.cache().get_value(cache_key)
    
    if not cached or cached.get("status") != "finished":
        frappe.throw("Archivo no disponible", frappe.PermissionError)

    file_url = cached.get("file_url")

    file_path = frappe.get_site_path("private", "files", file_url.split("/private/files/")[-1])

    if not os.path.exists(file_path):
        frappe.throw("Archivo no encontrado en disco")

    with open(file_path, "rb") as f:
        content = f.read()

    fname = file_url.split("/")[-1]
    frappe.local.response.filename = fname
    frappe.local.response.filecontent = content
    frappe.local.response.type = "download"
    frappe.local.response.display_content_as = "attachment"


def _render_report_pdf(doctype_name, report_row, print_format):
    html = frappe.get_print(
        doctype_name,
        report_row.name,
        print_format=print_format,
        as_pdf=False,
        no_letterhead=1,
    )
    html = _ensure_pdf_header_footer_placeholders(html)
    html = _compile_css_for_pdf(html)
    return get_pdf(html, options=_leadership_pdf_options())


def _resolve_file_name(survey, doctype_name, report_row):
    if not survey.su_is_leadership:
        label = report_row.get("cutoff_name") or "cutoff"
    else:
        label = report_row.get("leader_name") or report_row.name
    return f"{_sanitize_filename(label)}_{report_row.name}"