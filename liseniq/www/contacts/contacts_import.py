import frappe
from frappe import _
import io, csv
from datetime import datetime
import re
import json
import random
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Definición de columnas base
STANDARD_COLUMNS = [
	"Nombre", "Apellido", "Tipo de Documento", "Número de Documento (DNI)",
	"País", "Idioma", "Estatus", "Género", 
	"Fecha de Nacimiento", "Nivel Académico", "Correo (Opcional)", 
	"Fecha de Ingreso"
]

# Campos que no pueden estar vacíos
MANDATORY_FIELDS = [
	"Nombre", 
	"Apellido", 
	"Tipo de Documento", 
	"Número de Documento (DNI)", 
	"País", 
	"Idioma"
]

CHILD_TABLE_FIELD = "custom_additional_details"

# Lista fija de países LATAM
LATAM_COUNTRIES = [
	"Argentina", "Bolivia", "Brazil", "Chile", "Colombia", "Costa Rica",
	"Cuba", "Dominican Republic", "Ecuador", "El Salvador", "Guatemala",
	"Honduras", "Mexico", "Nicaragua", "Panama", "Paraguay", "Peru",
	"Puerto Rico", "Uruguay", "Venezuela, Bolivarian Republic of"
]

def get_context(context):
	# Validar si hay un proceso de carga masiva activo para la empresa
	try:
		user = frappe.session.user
		contact_info = frappe.db.get_value("Contact", {"user": user}, ["custom_company"], as_dict=True)
		
		if contact_info and contact_info.custom_company:
			# Verificar logs con estado Pendiente o Procesando
			is_active = frappe.db.exists("qp_IQ_UploadLog", {
				"ul_company": contact_info.custom_company,
				"ul_status": ["in", ["Pendiente", "Procesando"]]
			})
			
			if is_active:
				# Redireccionar a la lista de contactos si hay proceso activo
				frappe.local.response["type"] = "redirect"
				frappe.local.response["location"] = "/contacts"
				return
	except Exception:
		frappe.log_error(frappe.get_traceback(), "Error validando proceso activo en Carga Masiva")

	context.page_title = _("Carga masiva de Contactos")
	context.no_cache = 1
	context.no_breadcrumbs = True
	context.is_navbar_custom = True
	return context

def get_all_demographic_types():
	"""Retorna una lista de títulos de demográficos existentes en el sistema."""
	return [d.dt_title for d in frappe.get_all("qp_IQ_DemographicType", filters={"dt_object_type": "Contacto"}, fields=["dt_title"], order_by="dt_title asc")]

def get_mapping_dicts():
	"""
	Retorna diccionarios para mapear ID -> Nombre (para exportar) 
	y Nombre -> ID (para importar).
	"""
	# Tipo de Documento
	doc_types = frappe.get_all("qp_IQ_DocumentType", fields=["name", "dt_name"])
	dt_id_to_name = {d.name: d.dt_name for d in doc_types}
	dt_name_to_id = {d.dt_name: d.name for d in doc_types}

	# Idioma
	langs = frappe.get_all("qp_IQ_Language", fields=["name", "la_name"])
	lang_id_to_name = {d.name: d.la_name for d in langs}
	lang_name_to_id = {d.la_name: d.name for d in langs}

	# País
	countries = frappe.get_all("Country", fields=["name", "country_name"])
	country_id_to_name = {d.name: d.country_name for d in countries}
	country_name_to_id = {d.country_name: d.name for d in countries}

	# Nivel Académico
	academics = frappe.get_all("qp_IQ_AcademicLevel", fields=["name", "al_title"])
	academic_id_to_name = {d.name: d.al_title for d in academics}
	academic_name_to_id = {d.al_title: d.name for d in academics}
	
	# Demográficos: dt_title (Visible) <-> name (ID)
	demos = frappe.get_all("qp_IQ_DemographicType", fields=["name", "dt_title"])
	demo_title_to_id = {d.dt_title: d.name for d in demos}

	return {
		"dt_export": dt_id_to_name, "dt_import": dt_name_to_id,
		"lang_export": lang_id_to_name, "lang_import": lang_name_to_id,
		"country_export": country_id_to_name, "country_import": country_name_to_id,
		"academic_export": academic_id_to_name, "academic_import": academic_name_to_id,
		"demo_import": demo_title_to_id
	}

@frappe.whitelist()
def get_grid_options():
	"""
	Retorna las opciones para selectores y la lista de columnas dinámicas de demográficos.
	"""
	try:
		return {
			"document_types": [d.dt_name for d in frappe.get_all("qp_IQ_DocumentType", fields=["dt_name"], order_by="dt_name asc")],
			"languages": [d.la_name for d in frappe.get_all("qp_IQ_Language", fields=["la_name"], order_by="la_name asc")],
			"countries": sorted(LATAM_COUNTRIES),
			"genders": [d.gender for d in frappe.get_all("Gender", fields=["gender"], order_by="gender asc")],
			"academic_levels": [d.al_title for d in frappe.get_all("qp_IQ_AcademicLevel", fields=["al_title"], order_by="al_title asc")],
			"status": ["Activo", "Inactivo"],
			"demographic_headers": get_all_demographic_types() # Lista de columnas dinámicas
		}
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Error obteniendo opciones para Grid")
		return {}

def find_or_create_demographic_type(demographic_title):
	"""
	Busca el ID de un tipo demográfico por su título. Si no existe, lo crea.
	"""
	normalized_title = " ".join(demographic_title.strip().split()).title()
	object_type = "Contacto"

	if not normalized_title:
		return None

	try:
		existing_doc_name = frappe.db.get_value(
			"qp_IQ_DemographicType",
			{"dt_title": normalized_title, "dt_object_type": object_type},
			"name"
		)

		if existing_doc_name:
			return existing_doc_name
		else:
			doc = frappe.new_doc("qp_IQ_DemographicType")
			doc.dt_title = normalized_title
			doc.dt_object_type = object_type
			doc.dt_tag_color = "#{:06x}".format(random.randint(0, 0xFFFFFF))
			doc.dt_description = _("Demográfico '{0}' creado automáticamente desde Carga Masiva.").format(normalized_title)
			doc.insert(ignore_permissions=True)
			return doc.name

	except Exception:
		# Recuperación en caso de concurrencia
		return frappe.db.get_value(
			"qp_IQ_DemographicType",
			{"dt_title": normalized_title, "dt_object_type": object_type},
			"name"
		)

def _get_contacts_data_internal(company):
	"""
	Función auxiliar interna para obtener los datos de contactos formateados para el grid.
	Se reutiliza en get_contacts_for_grid y validate_contacts.
	"""
	contacts = frappe.get_all("Contact", 
		filters={
			"custom_company": company, 
			"custom_is_liseniq_contact": 1,
			"custom_is_deleted": 0
		},
		fields=[
			"name", "first_name", "last_name", "gender", "custom_dob",
			"custom_country", "custom_document_type", "custom_document_number",
			"custom_academic_level", "custom_entry_date", "custom_status",
			"custom_language"
		],
		order_by="first_name asc"
	)

	maps = get_mapping_dicts()
	demographic_headers = get_all_demographic_types()
	grid_rows = []
	
	for c in contacts:
		email = frappe.db.get_value("Contact Email", {"parent": c.name, "is_primary": 1}, "email_id")
		
		# Obtener demográficos
		demographics = frappe.get_all("qp_IQ_ContactAdditionalDetail", 
			filters={"parent": c.name},
			fields=["cad_tag", "cad_value"]
		)
		
		tipo_doc_label = maps["dt_export"].get(c.custom_document_type, c.custom_document_type)
		pais_label = maps["country_export"].get(c.custom_country, c.custom_country)
		idioma_label = maps["lang_export"].get(c.custom_language, c.custom_language)
		academic_label = maps["academic_export"].get(c.custom_academic_level, c.custom_academic_level)

		row = {
			"Nombre": c.first_name,
			"Apellido": c.last_name,
			"Tipo de Documento": tipo_doc_label or "",
			"Número de Documento (DNI)": c.custom_document_number,
			"País": pais_label or "",
			"Idioma": idioma_label or "",
			"Estatus": c.custom_status,
			"Género": c.gender,
			"Fecha de Nacimiento": str(c.custom_dob) if c.custom_dob else "",
			"Nivel Académico": academic_label or "",
			"Correo (Opcional)": email or "",
			"Fecha de Ingreso": str(c.custom_entry_date) if c.custom_entry_date else ""
		}
		
		# Llenar columnas dinámicas
		for demo in demographics:
			if demo.cad_tag:
				row[demo.cad_tag] = demo.cad_value

		grid_rows.append(row)

	return {"rows": grid_rows, "demographic_headers": demographic_headers}

@frappe.whitelist(allow_guest=False)
def get_contacts_for_grid():
	try:
		user = frappe.session.user
		contact_info = frappe.db.get_value(
			"Contact", 
			{"user": user, "custom_is_liseniq_contact": 0}, 
			["name", "custom_company"], 
			as_dict=True
		)
		
		if not contact_info or not contact_info.custom_company:
			return {"rows": [], "demographic_headers": []}

		return _get_contacts_data_internal(contact_info.custom_company)

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "FATAL ERROR: get_contacts_for_grid")
		return {"rows": [], "demographic_headers": []}

@frappe.whitelist(allow_guest=False)
def download_template():
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font
		
		# Obtener contexto del usuario y compañía primero
		user = frappe.session.user
		contact_info = frappe.db.get_value(
			"Contact", 
			{"user": user, "custom_is_liseniq_contact": 0}, 
			["name", "custom_company"], 
			as_dict=True
		)

		company = contact_info.custom_company if contact_info else None
		demographic_headers = []

		if company:
			# Obtener headers demográficos solo de los contactos existentes y activos de la empresa
			used_tags = frappe.db.sql("""
				SELECT DISTINCT d.cad_tag
				FROM `tabqp_IQ_ContactAdditionalDetail` d
				INNER JOIN `tabContact` c ON d.parent = c.name
				WHERE c.custom_company = %s
				AND c.custom_is_liseniq_contact = 1
				AND c.custom_is_deleted = 0
				AND d.cad_tag IS NOT NULL AND d.cad_tag != ''
				ORDER BY d.cad_tag ASC
			""", (company,), as_dict=True)
			
			demographic_headers = [d.cad_tag for d in used_tags]

		wb = Workbook()
		ws = wb.active
		ws.title = "Plantilla Contactos"

		ws_opts = wb.create_sheet("Opciones")
		ws_opts.sheet_state = 'hidden' 

		doc_types_list = [d.dt_name for d in frappe.get_all("qp_IQ_DocumentType", fields=["dt_name"], order_by="dt_name asc")]
		langs_list = [d.la_name for d in frappe.get_all("qp_IQ_Language", fields=["la_name"], order_by="la_name asc")]
		genders_list = [d.gender for d in frappe.get_all("Gender", fields=["gender"], order_by="gender asc")]
		countries_list = sorted([c for c in LATAM_COUNTRIES])
		status_list = ["Activo", "Inactivo"]
		academic_levels_list = [d.al_title for d in frappe.get_all("qp_IQ_AcademicLevel", fields=["al_title"], order_by="al_title asc")]

		lists_map = [
			(doc_types_list, "A"),    
			(countries_list, "B"),    
			(langs_list, "C"),        
			(genders_list, "D"),      
			(status_list, "E"),       
			(academic_levels_list, "F") 
		]

		for data_list, col_letter in lists_map:
			for idx, val in enumerate(data_list, start=1):
				ws_opts[f"{col_letter}{idx}"] = val

		# Construir Headers: Estándar + Dinámicos Filtrados
		headers = list(STANDARD_COLUMNS) + demographic_headers
		
		ws.append(headers)
		for cell in ws[1]:
			cell.font = Font(bold=True)

		def create_dv(formula):
			dv = DataValidation(type="list", formula1=formula, allow_blank=True)
			return dv

		dv_doctype = create_dv(f"'Opciones'!$A$1:$A${len(doc_types_list) or 1}")
		dv_country = create_dv(f"'Opciones'!$B$1:$B${len(countries_list) or 1}")
		dv_lang    = create_dv(f"'Opciones'!$C$1:$C${len(langs_list) or 1}")
		dv_gender  = create_dv(f"'Opciones'!$D$1:$D${len(genders_list) or 1}")
		dv_status  = create_dv(f"'Opciones'!$E$1:$E${len(status_list) or 1}")
		dv_academic = create_dv(f"'Opciones'!$F$1:$F${len(academic_levels_list) or 1}")

		ws.add_data_validation(dv_doctype)
		ws.add_data_validation(dv_country)
		ws.add_data_validation(dv_lang)
		ws.add_data_validation(dv_gender)
		ws.add_data_validation(dv_status)
		ws.add_data_validation(dv_academic)

		col_map = {name: i+1 for i, name in enumerate(headers)}
		start_row = 2
		end_row = 5000 # Rango extendido para data existente

		if "Tipo de Documento" in col_map: dv_doctype.add(f"{get_column_letter(col_map['Tipo de Documento'])}{start_row}:{get_column_letter(col_map['Tipo de Documento'])}{end_row}")
		if "País" in col_map: dv_country.add(f"{get_column_letter(col_map['País'])}{start_row}:{get_column_letter(col_map['País'])}{end_row}")
		if "Idioma" in col_map: dv_lang.add(f"{get_column_letter(col_map['Idioma'])}{start_row}:{get_column_letter(col_map['Idioma'])}{end_row}")
		if "Género" in col_map: dv_gender.add(f"{get_column_letter(col_map['Género'])}{start_row}:{get_column_letter(col_map['Género'])}{end_row}")
		if "Estatus" in col_map: dv_status.add(f"{get_column_letter(col_map['Estatus'])}{start_row}:{get_column_letter(col_map['Estatus'])}{end_row}")
		if "Nivel Académico" in col_map: dv_academic.add(f"{get_column_letter(col_map['Nivel Académico'])}{start_row}:{get_column_letter(col_map['Nivel Académico'])}{end_row}")

		if company:
			data = _get_contacts_data_internal(company)
			
			for row_dict in data['rows']:
				row_values = []
				for h in headers:
					row_values.append(row_dict.get(h, ""))
				
				ws.append(row_values)

		output = io.BytesIO()
		wb.save(output)
		output.seek(0)
		
		timestamp = frappe.utils.now_datetime().strftime("%Y%m%d_%H%M")
		frappe.local.response['filename'] = f"Plantilla_Contactos_{timestamp}.xlsx"
		frappe.local.response["filecontent"] = output.read()
		frappe.local.response["type"] = "download"

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "FATAL ERROR: download_template")
		frappe.throw(_("No se pudo generar la plantilla: {0}").format(e))

def check_if_modified(contact_doc, data, status):
	"""
	Compara el documento actual con los nuevos datos para ver si hay cambios reales.
	"""
	def norm(val): return str(val).strip() if val is not None else ""
	def norm_date(val): return str(val) if val else None

	# 1. Campos directos
	if norm(contact_doc.first_name) != norm(data['firstName']): return True
	if norm(contact_doc.last_name) != norm(data['lastName']): return True
	if norm(contact_doc.custom_document_type) != norm(data['docType']): return True
	if norm(contact_doc.custom_country) != norm(data['country']): return True
	if norm(contact_doc.custom_language) != norm(data['language']): return True
	if norm(contact_doc.custom_academic_level) != norm(data['education']): return True
	if norm(contact_doc.gender) != norm(data['gender']): return True
	if norm(contact_doc.custom_status) != norm(status): return True

	# 2. Fechas
	if norm_date(contact_doc.custom_dob) != str(data['birthdate']) if data['birthdate'] else None: return True
	if norm_date(contact_doc.custom_entry_date) != str(data['entryDate']) if data['entryDate'] else None: return True

	# 3. Email
	current_email = frappe.db.get_value("Contact Email", {"parent": contact_doc.name, "is_primary": 1}, "email_id")
	if norm(current_email) != norm(data['email']): return True

	# 4. Demográficos
	current_demos = frappe.get_all("qp_IQ_ContactAdditionalDetail", filters={"parent": contact_doc.name}, fields=["cad_tag", "cad_value"])
	current_set = set((norm(d.cad_tag), norm(d.cad_value)) for d in current_demos)
	new_set = set((norm(d['type']), norm(d['value'])) for d in data['demographics'])
	
	if current_set != new_set: return True

	return False

def update_contact_fields(contact_doc, data, status, demo_map=None):
	"""Actualiza los campos del documento en memoria y guarda."""
	contact_doc.first_name = data['firstName']
	contact_doc.last_name = data['lastName']
	contact_doc.custom_document_type = data['docType']
	contact_doc.custom_country = data['country']
	contact_doc.custom_language = data['language']
	contact_doc.custom_academic_level = data['education']
	contact_doc.gender = data['gender']
	contact_doc.custom_dob = data['birthdate']
	contact_doc.custom_entry_date = data['entryDate']
	contact_doc.custom_status = status
	
	contact_doc.save(ignore_permissions=True)
	
	new_email = (data.get('email') or "").strip()
	if new_email:
		exists = False
		for email_row in contact_doc.email_ids:
			if email_row.is_primary:
				email_row.email_id = new_email
				exists = True
				break
		if not exists:
			contact_doc.append("email_ids", {"email_id": new_email, "is_primary": 1})
		contact_doc.save(ignore_permissions=True)
	
	# Validar existencia del campo de tabla hija en metadata
	if not contact_doc.meta.get_field(CHILD_TABLE_FIELD):
		frappe.log_error(f"Error Crítico: El campo '{CHILD_TABLE_FIELD}' no existe en el DocType Contact.", "Import Contacts Error")
		return

	# Reemplazar demográficos
	contact_doc.set(CHILD_TABLE_FIELD, [])
	
	for d in data['demographics']:
		if d['type'] and d['value']:
			demo_id = None
			if demo_map:
				demo_id = demo_map.get(d['type'])
			
			if not demo_id:
				demo_id = find_or_create_demographic_type(d['type'])
			
			child = contact_doc.append(CHILD_TABLE_FIELD, {})
			child.cad_demographic_type = demo_id
			child.cad_tag = d['type']
			child.cad_value = d['value']
	
	contact_doc.save(ignore_permissions=True)

def create_upload_log(file_name, total_rows, user, company):
	"""
	Crea el registro de Upload Log con estatus 'Pendiente'
	"""
	log = frappe.new_doc("qp_IQ_UploadLog")
	log.ul_file_name = file_name
	log.ul_status = "Pendiente"
	log.ul_total_rows = total_rows
	log.ul_processed_rows = 0
	log.ul_success_count = 0
	log.ul_error_count = 0
	log.ul_owner = user
	log.ul_company = company
	log.insert(ignore_permissions=True)
	return log.name

def process_contacts_background(log_name, rows, user):
	"""
	Procesa los contactos en segundo plano.
	Actualiza el registro qp_IQ_UploadLog a medida que avanza.
	"""
	try:
		# Obtener log doc
		log_doc = frappe.get_doc("qp_IQ_UploadLog", log_name)
		log_doc.ul_status = "Procesando"
		log_doc.ul_started_at = frappe.utils.now()
		log_doc.ul_error_log = "[]" # Inicializar como lista vacía JSON
		log_doc.save(ignore_permissions=True)
		frappe.db.commit()

		contact_info = frappe.db.get_value("Contact", {"user": user, "custom_is_liseniq_contact": 0}, ["name", "custom_company"], as_dict=True)
		if not contact_info or not contact_info.custom_company:
			log_doc.ul_status = "Fallido"
			log_doc.ul_completed_at = frappe.utils.now()
			log_doc.ul_error_log = json.dumps([{"fila": 0, "error": "No se pudo determinar la compañía del usuario"}])
			log_doc.save(ignore_permissions=True)
			return

		user_company = contact_info.custom_company

		existing_contacts_map = {}
		active_contacts = frappe.get_all("Contact", 
			filters={
				"custom_company": user_company, 
				"custom_is_liseniq_contact": 1,
				"custom_is_deleted": 0 
			}, 
			fields=["name", "custom_document_number"]
		)
		for c in active_contacts:
			if c.custom_document_number:
				existing_contacts_map[str(c.custom_document_number).strip()] = c.name
		
		processed_dnis_in_file = set()

		# Obtener mapas (incluyendo el nuevo mapa de demográficos)
		maps = get_mapping_dicts()
		dt_map = maps["dt_import"]
		country_map = maps["country_import"]
		lang_map = maps["lang_import"]
		academic_map = maps["academic_import"]
		demo_map = maps["demo_import"]

		def parse_date(value):
			if not value: return None
			try: return frappe.utils.getdate(value)
			except:
				try: return datetime.strptime(value, "%d/%m/%Y").date()
				except: return None

		success_count = 0
		error_count = 0
		processed_count = 0
		error_list = []

		for i, r in enumerate(rows, start=1):
			try:
				# Extracción y limpieza de datos básicos
				tipo_doc_raw = (r.get("Tipo de Documento") or "").strip()
				tipo_doc_id = dt_map.get(tipo_doc_raw, tipo_doc_raw)
				pais_raw = (r.get("País") or "").strip()
				pais_id = country_map.get(pais_raw, pais_raw)
				idioma_raw = (r.get("Idioma") or "").strip()
				idioma_id = lang_map.get(idioma_raw, idioma_raw)
				nivel_acad_raw = (r.get("Nivel Académico") or "").strip()
				nivel_acad_id = academic_map.get(nivel_acad_raw, nivel_acad_raw)
				
				nombre = (r.get("Nombre") or "").strip()
				apellido = (r.get("Apellido") or "").strip()
				numero_doc = (r.get("Número de Documento (DNI)") or "").strip()
				estatus = (r.get("Estatus") or "").strip()
				correo = (r.get("Correo (Opcional)") or "").strip()

				if numero_doc:
					processed_dnis_in_file.add(numero_doc)

				fecha_nac = parse_date(r.get("Fecha de Nacimiento"))
				fecha_ing = parse_date(r.get("Fecha de Ingreso"))
				
				# Validación de campos obligatorios
				missing = []
				if not nombre: missing.append("Nombre")
				if not apellido: missing.append("Apellido")
				if not tipo_doc_id: missing.append("Tipo de Documento")
				if not numero_doc: missing.append("Número de Documento")
				if not pais_id: missing.append("País")
				if not idioma_id: missing.append("Idioma")
				
				if missing:
					raise Exception(f"Campos obligatorios faltantes: {', '.join(missing)}")

				data = {
					"firstName": nombre,
					"lastName": apellido,
					"docNumber": numero_doc,
					"docType": tipo_doc_id,
					"country": pais_id,
					"language": idioma_id,
					"email": correo,
					"gender": (r.get("Género") or "").strip(),
					"education": nivel_acad_id,
					"birthdate": fecha_nac,
					"entryDate": fecha_ing,
					"demographics": []
				}
				
				# Recolectar demográficos dinámicos
				for col_name, val in r.items():
					clean_col = col_name.strip()
					clean_val = str(val or "").strip()
					
					if clean_col not in STANDARD_COLUMNS and clean_col and clean_val:
						data["demographics"].append({"type": clean_col, "value": clean_val})

				# Verificar si existe
				contact_name = None
				if numero_doc:
					contact_name = frappe.db.get_value("Contact", {"custom_document_number": numero_doc, "custom_company": user_company}, "name")

				if contact_name:
					contact_doc = frappe.get_doc("Contact", contact_name)

					if contact_doc.custom_is_deleted:
						contact_doc.custom_is_deleted = 0
						contact_doc.save(ignore_permissions=True)
						
					if check_if_modified(contact_doc, data, estatus or contact_doc.custom_status):
						update_contact_fields(contact_doc, data, estatus or contact_doc.custom_status, demo_map)
				else:
					new_doc = frappe.new_doc("Contact")
					new_doc.first_name = nombre
					new_doc.last_name = apellido
					new_doc.custom_company = user_company
					new_doc.custom_document_number = numero_doc
					new_doc.custom_document_type = tipo_doc_id
					new_doc.custom_country = pais_id
					new_doc.custom_language = idioma_id
					new_doc.custom_academic_level = nivel_acad_id
					new_doc.custom_status = estatus or "Activo"
					new_doc.custom_is_liseniq_contact = 1
					new_doc.custom_is_deleted = 0
					
					new_doc.custom_dob = data['birthdate']
					new_doc.custom_entry_date = data['entryDate']
					new_doc.gender = data['gender']

					if correo:
						new_doc.append("email_ids", {"email_id": correo, "is_primary": 1})
					
					new_doc.insert(ignore_permissions=True)
					
					if data['demographics']:
						if new_doc.meta.get_field(CHILD_TABLE_FIELD):
							for d in data['demographics']:
								demo_id = demo_map.get(d['type'])
								if not demo_id:
									demo_id = find_or_create_demographic_type(d['type'])
								
								child = new_doc.append(CHILD_TABLE_FIELD, {})
								child.cad_demographic_type = demo_id
								child.cad_tag = d['type']
								child.cad_value = d['value']
							
							new_doc.save(ignore_permissions=True)
				
				success_count += 1

			except Exception as e:
				error_count += 1
				# Agregar al log de errores JSON
				error_entry = {"fila": i, "error": str(e)}
				error_list.append(error_entry)

			# Actualizar progreso periódicamente (cada 5 registros) para no saturar DB
			processed_count += 1
			if processed_count % 5 == 0:
				log_doc = frappe.get_doc("qp_IQ_UploadLog", log_name)
				log_doc.ul_processed_rows = processed_count
				log_doc.ul_success_count = success_count
				log_doc.ul_error_count = error_count
				# Actualizar JSON de errores incrementalmente si es necesario, 
				# pero por eficiencia lo guardamos completo o en chunks. 
				# Aquí guardamos el estado actual.
				log_doc.ul_error_log = json.dumps(error_list)
				log_doc.save(ignore_permissions=True)
				frappe.db.commit()

		delete_count = 0
		try:
			existing_dnis_set = set(existing_contacts_map.keys())
			missing_dnis = existing_dnis_set - processed_dnis_in_file
			
			for missing_dni in missing_dnis:
				contact_name_to_delete = existing_contacts_map[missing_dni]
				frappe.db.set_value("Contact", contact_name_to_delete, "custom_is_deleted", 1)
				delete_count += 1
				
		except Exception as e:
			# Loguear error de eliminación pero no detener el proceso general si ya se procesaron filas
			error_list.append({"fila": "N/A", "error": f"Error en proceso de eliminación lógica: {str(e)}"})

		# Finalización
		log_doc = frappe.get_doc("qp_IQ_UploadLog", log_name)
		log_doc.ul_processed_rows = processed_count
		log_doc.ul_success_count = success_count
		log_doc.ul_error_count = error_count
		log_doc.ul_completed_at = frappe.utils.now()
		log_doc.ul_error_log = json.dumps(error_list)

		if error_count == 0:
			log_doc.ul_status = "Completado"
		elif success_count == 0 and error_count > 0:
			log_doc.ul_status = "Fallido"
		else:
			log_doc.ul_status = "Completado con errores"
		
		log_doc.save(ignore_permissions=True)

		# Crear Notificación de Portal
		try:
			notification = frappe.new_doc("qp_IQ_PortalNotification")
			notification.pn_user = user
			notification.pn_title = "Carga Masiva de Contactos"
			notification.pn_message = f"Carga finalizada ({log_doc.ul_status})\n\n✅ Exitosos: {success_count}\n❌ Fallidos: {error_count}\n🗑️ Contactos eliminados: {delete_count}"
			
			notification.pn_route = "/contacts"
			notification.pn_type = "Info" if error_count == 0 else "Warning"
			notification.pn_is_read = 0
			notification.insert(ignore_permissions=True)
		except Exception as e:
			frappe.log_error(f"Error creando notificación de portal: {str(e)}", "Portal Notification Error")

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "FATAL ERROR: process_contacts_background")
		# Intentar marcar como fallido el log si algo catastrófico ocurre
		try:
			log_doc = frappe.get_doc("qp_IQ_UploadLog", log_name)
			log_doc.ul_status = "Fallido"
			log_doc.ul_error_log = json.dumps([{"fila": 0, "error": f"Error fatal de sistema: {str(e)}"}])
			log_doc.save(ignore_permissions=True)
		except:
			pass

@frappe.whitelist(allow_guest=False)
def upload_contacts():
	if not frappe.request: frappe.throw(_("No hay request disponible"))
	fileobj = frappe.local.request.files.get('file')
	if not fileobj: frappe.throw(_("No se envió ningún archivo"))

	filename = fileobj.filename or "Carga_Archivo.xlsx"
	ext = filename.split('.')[-1].lower()
	rows = []
	try:
		if ext in ("xlsx", "xls"):
			from openpyxl import load_workbook
			wb = load_workbook(fileobj, read_only=True, data_only=True)
			if "Plantilla Contactos" in wb.sheetnames: ws = wb["Plantilla Contactos"]
			else: ws = wb.active
			for r in ws.iter_rows(values_only=True):
				rows.append([None if v is None else str(v).strip() for v in r])
		elif ext == "csv":
			content = fileobj.read()
			if isinstance(content, bytes): content = content.decode('utf-8-sig')
			reader = csv.reader(io.StringIO(content))
			for r in reader:
				rows.append([None if v is None else v.strip() for v in r])
		else: frappe.throw(_("Tipo de archivo no soportado"))
	except Exception as e: frappe.throw(_("Error al leer archivo: {0}").format(e))

	if not rows or len(rows) < 1: frappe.throw(_("Archivo vacío"))

	headers = [h.strip() if h else "" for h in rows[0]]
	idx = {h: i for i, h in enumerate(headers)}
	
	if "Nombre" not in idx or "Apellido" not in idx:
		frappe.throw(_("Faltan columnas obligatorias (Nombre, Apellido)"))

	def get_val(r, col):
		if col not in idx: return ""
		i = idx[col]
		return str(r[i]).strip() if i < len(r) and r[i] else ""
	
	rows_dicts = []
	for i, r in enumerate(rows[1:], start=2):
		if not any(cell for cell in r): continue
		row_dict = {}
		for h in headers:
			row_dict[h] = get_val(r, h)
		rows_dicts.append(row_dict)

	# Obtener compañía del usuario para el log
	user = frappe.session.user
	contact_info = frappe.db.get_value("Contact", {"user": user}, ["custom_company"], as_dict=True)
	company = contact_info.custom_company if contact_info else None

	if not company:
		frappe.throw(_("No se pudo determinar la compañía del usuario"))

	# Crear Log
	log_name = create_upload_log(filename, len(rows_dicts), user, company)

	# Encolar proceso pasando el log_name
	frappe.enqueue(
		method=process_contacts_background,
		queue='long',
		timeout=3600,
		log_name=log_name,
		rows=rows_dicts,
		user=user
	)
	
	return {"message": {"total_rows": len(rows_dicts), "queued": True, "log_name": log_name}, "status": "queued"}

@frappe.whitelist(allow_guest=False)
def validate_contacts():
	if not frappe.request: frappe.throw(_("No hay request disponible"))
	fileobj = frappe.local.request.files.get('file')
	if not fileobj: frappe.throw(_("Error de validación: No se detectó ningún archivo adjunto."))

	# Obtener información del usuario para validaciones contra DB
	user = frappe.session.user
	contact_info = frappe.db.get_value("Contact", {"user": user}, ["custom_company"], as_dict=True)
	company = contact_info.custom_company if contact_info else None
	
	existing_grid_rows = []
	if company:
		# Obtengo TODOS los datos de contactos, formateados igual que para el grid
		# Lo uso para en el frontend comparar campo a campo para detectar cambios reales
		internal_data = _get_contacts_data_internal(company)
		existing_grid_rows = internal_data.get("rows", [])

	# Obtener opciones válidas para listas
	options = get_grid_options()

	filename = fileobj.filename or ""
	ext = filename.split('.')[-1].lower()
	rows = []
	try:
		if ext in ("xlsx", "xls"):
			from openpyxl import load_workbook
			wb = load_workbook(fileobj, read_only=True, data_only=True)
			if "Plantilla Contactos" in wb.sheetnames: ws = wb["Plantilla Contactos"]
			else: ws = wb.active
			for r in ws.iter_rows(values_only=True):
				rows.append([None if v is None else str(v).strip() for v in r])
		elif ext == "csv":
			content = fileobj.read()
			if isinstance(content, bytes): content = content.decode('utf-8-sig')
			reader = csv.reader(io.StringIO(content))
			for r in reader:
				rows.append([None if v is None else v.strip() for v in r])
	except Exception as e: return {"ok": False, "error": f"Error leyendo archivo: {str(e)}"}

	if not rows: return {"ok": False, "error": "El archivo está vacío"}

	# Validaciones preliminares
	# Validar que la celda A1 (Columna 1, Fila 1) no esté vacía
	if not rows[0] or not rows[0][0] or not str(rows[0][0]).strip():
		return {"ok": False, "error": _("La celda ubicada en la columna 1, fila 1 no puede estar vacía.")}

	headers = [str(h).strip() if h else "" for h in rows[0]]

	# Validar que los demográficos tengan nombre
	if "Fecha de Ingreso" in headers:
		fi_index = headers.index("Fecha de Ingreso")
		for i in range(fi_index + 1, len(headers)):
			if not headers[i]:
				return {"ok": False, "error": _("Columna sin nombre en la posición {0} - Los campos demográficos deben tener un encabezado válido.").format(i+1)}

	idx = {h: i for i, h in enumerate(headers)}
	parsed = []
	errors = []

	def get_val(r, col):
		if col not in idx: return ""
		i = idx[col]
		return str(r[i]).strip() if i < len(r) and r[i] else ""
	
	# Contadores de reporte preliminar
	stats = {
		"new": 0,
		"existing": 0,
		"errors": 0
	}

	for i, r in enumerate(rows[1:], start=2):
		if not any(cell for cell in r): continue
		row_dict = {}
		for h in headers:
			val = get_val(r, h)
			row_dict[h] = val
		
		row_errors = []
		
		# Campos Obligatorios
		for field in MANDATORY_FIELDS:
			if not row_dict.get(field):
				row_errors.append(f"Falta {field}")

		# Validación de Listas (Selects)
		field_map = {
			"Tipo de Documento": "document_types",
			"País": "countries",
			"Idioma": "languages",
			"Estatus": "status",
			"Género": "genders",
			"Nivel Académico": "academic_levels"
		}
		
		for field_name, option_key in field_map.items():
			val = row_dict.get(field_name)
			if val and option_key in options:
				if val not in options[option_key]:
					row_errors.append(f"'{val}' no es válido para {field_name}")
		
		if row_errors:
			stats["errors"] += 1
			errors.append({"fila": i, "errores": row_errors})
		
		parsed.append(row_dict)

	# Retornamos existing_grid_rows con todos los datos en lugar de solo DNIs
	return {
		"ok": True, 
		"headers": headers, 
		"rows": parsed, 
		"errors": errors,
		"stats": stats,
		"existing_grid_rows": existing_grid_rows,
		"valid_options": options
	}

@frappe.whitelist(allow_guest=False)
def upload_contacts_json(rows_json, file_name=None):
	import json as _json
	try: rows = _json.loads(rows_json)
	except: frappe.throw(_("JSON inválido"))

	# Obtener compañía
	user = frappe.session.user
	contact_info = frappe.db.get_value("Contact", {"user": user}, ["custom_company"], as_dict=True)
	company = contact_info.custom_company if contact_info else None

	if not company:
		frappe.throw(_("No se pudo determinar la compañía del usuario"))

	# Determinar nombre del archivo
	final_file_name = file_name or "Edición en Linea"

	# Crear Log con nombre del archivo recibido o "Edición en Linea"
	log_name = create_upload_log(final_file_name, len(rows), user, company)

	# Encolar proceso
	frappe.enqueue(
		method=process_contacts_background,
		queue='long',
		timeout=3600,
		log_name=log_name,
		rows=rows,
		user=user
	)
	
	return {"message": {"total_rows": len(rows), "queued": True, "log_name": log_name}, "status": "queued"}

@frappe.whitelist(allow_guest=False)
def check_upload_status():
	"""
	Verifica el estado de la carga más reciente de la compañía,
	sea activa o ya finalizada.
	"""
	user = frappe.session.user
	contact_info = frappe.db.get_value("Contact", {"user": user}, ["custom_company"], as_dict=True)
	if not contact_info or not contact_info.custom_company:
		return {"active": False}

	company = contact_info.custom_company

	# Obtener el último log de carga de esta compañía (por fecha de creación descendente)
	last_log = frappe.get_all(
		"qp_IQ_UploadLog",
		filters={"ul_company": company},
		fields=["name", "ul_status", "ul_processed_rows", "ul_total_rows", "ul_file_name", "ul_success_count", "ul_error_count"],
		order_by="creation desc",
		limit=1
	)

	if not last_log:
		return {"active": False}

	log = last_log[0]
	is_active = log.ul_status in ["Pendiente", "Procesando"]

	return {
		"active": is_active,
		"status": log.ul_status,
		"processed": log.ul_processed_rows,
		"total": log.ul_total_rows,
		"success": log.ul_success_count,
		"error": log.ul_error_count,
		"file_name": log.ul_file_name
	}